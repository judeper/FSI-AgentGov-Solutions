#!/usr/bin/env python3
"""
Importer: Agent Registry export (XLSX or CSV) -> fsi_copilotagent owner enrichment.

The Microsoft 365 admin center (or Power Platform admin center) registry export
provides Owner and Date-created columns for agents not discoverable via API with
owner attribution (e.g., Agent Builder agents). This importer:

  1. Validates the source file against a configurable header alias map loaded from
     templates/registry-columnmap.sample.json (or a custom --columnmap path).
  2. Hard-fails with an explicit header-diff if required canonical columns are
     absent after alias mapping, so export-format drift surfaces immediately.
  3. Computes a SHA-256 of the source file for provenance.
  4. Tags every row:
       fsi_ownersource        = "Agent Registry Export"
       fsi_ownermatchconfidence = "Exact" when a stable owner object ID (Entra
                                  GUID, fsi_ownerid) is present in the row;
                                  "Heuristic" when only a UPN is present (no
                                  object ID); "Unmatched" when owner is blank
                                  or absent.
       fsi_ownerasofdatetime  = the --as-of value (ISO-8601 UTC) that the caller
                                 passes to signal the staleness of the export.

Supported formats:
  * XLSX — openpyxl (read_only=True); reads the first sheet by default or the
           sheet named by "sheet_name" in the column-map JSON.
  * CSV  — stdlib csv (utf-8-sig encoding, BOM-safe for M365 exports).
  Format is detected by file extension (.xlsx / .xls -> XLSX; anything else -> CSV).

Output: JSON array of row dicts written to --output or stdout.

# ---------------------------------------------------------------------------
# CONFIRMED HEADER CONTRACT — JULY 2026 LAB EXPORT
#   The validated Microsoft 365 admin center Registry workbook includes exact
#   headers "Name", "Bot Id", "Owner", "Date created", and "Creator Id".
#   Owner is UPN-shaped and maps to owner_upn. Creator Id is a distinct GUID and
#   must not be mapped to owner_id or otherwise treated as the owner.
#   The alias map remains configurable so export-contract drift fails visibly
#   without requiring code changes.
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# SOURCE-OF-TRUTH GUARDRAIL
#   This importer sources owner and creator attribution solely from the
#   supported, publicly documented Microsoft 365 admin center Registry export
#   (XLSX or CSV) or documented Microsoft Graph API endpoints.  It
#   intentionally does not consume undocumented or internal admin-center
#   backend routes, even where such routes may surface richer attribution
#   data — those routes do not constitute a supported public contract and must
#   not be used as the integration surface for this solution.
#
#   Confirmed native headers remain behind the configurable alias map
#   (templates/registry-columnmap.sample.json or a custom --columnmap path).
#   Any required canonical column that remains unmapped after alias resolution
#   is a hard failure with an explicit header-diff — silent or guessed mappings
#   are not permitted.
# ---------------------------------------------------------------------------
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column-map defaults
# ---------------------------------------------------------------------------

_SCRIPT_DIR = Path(__file__).parent
_DEFAULT_COLUMNMAP = _SCRIPT_DIR.parent / "templates" / "registry-columnmap.sample.json"

# Canonical field names produced after alias mapping. These are the internal
# names used by this importer; the alias map translates export-header variants
# to these names.
CANONICAL_AGENT_NAME = "agent_name"
CANONICAL_AGENT_ID = "agent_id"
CANONICAL_OWNER_UPN = "owner_upn"
CANONICAL_OWNER_ID = "owner_id"
CANONICAL_DATE_CREATED = "date_created"

# Default required canonical fields. Overridden by "required_canonical_fields"
# in the column-map JSON when present.
_DEFAULT_REQUIRED_FIELDS = frozenset({CANONICAL_AGENT_NAME, CANONICAL_OWNER_UPN})

# Mapping from canonical field name to Dataverse logical name.
CANONICAL_TO_DATAVERSE: dict[str, str] = {
    CANONICAL_AGENT_NAME: "fsi_agentname",
    CANONICAL_AGENT_ID:   "fsi_agentid",
    CANONICAL_OWNER_UPN:  "fsi_ownerupn",
    CANONICAL_OWNER_ID:   "fsi_ownerid",
    CANONICAL_DATE_CREATED: "fsi_createdon",
}

# Provenance tags stamped on every emitted row.
OWNER_SOURCE_LABEL = "Agent Registry Export"
_UNMAPPED_HEADER_PREFIX = "unmapped__"


# ---------------------------------------------------------------------------
# Column-map loading
# ---------------------------------------------------------------------------

def load_columnmap(path: Path) -> tuple[dict[str, str], frozenset[str], Optional[str]]:
    """Load the header alias map from a JSON file.

    Returns (header_aliases, required_canonical_fields, sheet_name).
    """
    if not path.exists():
        raise FileNotFoundError(
            f"Column-map file not found: {path}. "
            "Create or specify a column-map JSON (--columnmap)."
        )
    raw = json.loads(path.read_text(encoding="utf-8"))
    aliases: dict[str, str] = raw.get("header_aliases", {})
    required_raw: list = raw.get("required_canonical_fields", list(_DEFAULT_REQUIRED_FIELDS))
    required = frozenset(str(f) for f in required_raw)
    sheet_name: Optional[str] = raw.get("sheet_name") or None
    logger.debug(
        "Loaded column-map from %s: %d aliases, %d required fields, sheet=%r",
        path, len(aliases), len(required), sheet_name,
    )
    return aliases, required, sheet_name


def _normalize_header(header: str, aliases: dict[str, str]) -> str:
    """Map a raw export header to a canonical field name.

    Lookup is case-insensitive and strips leading/trailing whitespace.
    Exact canonical names remain valid for programmatic CSV inputs. Other
    unmapped headers are namespaced so a new display label such as "Owner Id"
    cannot silently become the canonical owner_id field.
    """
    key = header.strip().lower()
    mapped = aliases.get(key)
    if mapped:
        return mapped
    if key in CANONICAL_TO_DATAVERSE:
        return key
    return f"{_UNMAPPED_HEADER_PREFIX}{key.replace(' ', '_')}"


# ---------------------------------------------------------------------------
# SHA-256
# ---------------------------------------------------------------------------

def sha256_file(path: str) -> str:
    """Compute the SHA-256 hex digest of a file."""
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def _rows_from_xlsx(path: str, sheet_name: Optional[str]) -> list[list[Any]]:
    """Read rows from an XLSX file using openpyxl (read_only=True).

    Returns a list of lists (first row = headers).
    """
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to read XLSX files. "
            "Install via: pip install 'openpyxl>=3.1'"
        ) from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in {path}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active or wb.worksheets[0]

    all_rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))
    wb.close()
    return all_rows


def _rows_from_csv(path: str) -> list[list[str]]:
    """Read rows from a CSV file (utf-8-sig, BOM-safe)."""
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        return [list(row) for row in reader]


def _detect_format(path: str) -> str:
    """Return 'xlsx' or 'csv' based on file extension.

    Raises ValueError for legacy .xls files — the XLSX binary format produced
    by modern M365 admin center exports is .xlsx; .xls is an unsupported legacy
    format that openpyxl cannot safely read.
    """
    ext = Path(path).suffix.lower()
    if ext == ".xls":
        raise ValueError(
            f"Legacy .xls format is not supported: {path}. "
            "Export the registry as .xlsx or .csv from the Microsoft 365 admin center."
        )
    return "xlsx" if ext == ".xlsx" else "csv"


def _parse_iso8601_utc(value: str) -> str:
    """Parse and normalise an ISO-8601 UTC datetime string.

    Accepts common UTC forms: '2026-07-20T18:00:00Z', '2026-07-20T18:00:00+00:00',
    '2026-07-20T18:00:00.000Z'.  Returns the canonical 'YYYY-MM-DDTHH:MM:SSZ' form.

    Raises ValueError on invalid input or non-UTC timezone offsets.
    """
    s = value.strip()
    # Normalise 'Z' suffix to '+00:00' for fromisoformat() on Python 3.7-3.10.
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(s)
    except ValueError:
        raise ValueError(
            f"Invalid ISO-8601 datetime: {value!r}. "
            "Expected UTC, e.g. '2026-07-20T18:00:00Z'."
        )
    if dt.tzinfo is None:
        raise ValueError(
            f"Datetime {value!r} has no timezone — UTC is required (append 'Z')."
        )
    dt_utc = dt.astimezone(timezone.utc)
    return dt_utc.strftime("%Y-%m-%dT%H:%M:%SZ")


# ---------------------------------------------------------------------------
# Core import logic
# ---------------------------------------------------------------------------

def import_registry_file(
    path: str,
    aliases: dict[str, str],
    required_fields: frozenset[str],
    sheet_name: Optional[str],
    as_of: Optional[str],
) -> tuple[list[dict], list[str]]:
    """Parse the registry export and return (rows, warnings).

    Raises ValueError on:
      * .xls file extension (legacy format, unsupported — detected by _detect_format)
      * empty file
      * two distinct raw headers mapping to the same key canonical field
      * missing required canonical fields after alias mapping

    Per-row warnings (non-fatal) are collected in the returned list:
      * invalid ISO-8601 UTC values in date_created -> fsi_createdon set to null
    """
    fmt = _detect_format(path)  # raises ValueError for .xls
    logger.info("Detected format: %s for %s", fmt, path)

    if fmt == "xlsx":
        all_rows = _rows_from_xlsx(path, sheet_name)
    else:
        all_rows = _rows_from_csv(path)

    if not all_rows:
        raise ValueError(f"Registry export file is empty: {path}")

    # First row is the header.
    raw_headers: list[str] = [str(h or "").strip() for h in all_rows[0]]
    canonical_headers = [_normalize_header(h, aliases) for h in raw_headers]
    mapped_header_set = set(canonical_headers)

    # Reject ambiguous duplicate canonical header mappings for the five key fields.
    _DUPE_GUARD = {
        CANONICAL_AGENT_NAME, CANONICAL_AGENT_ID,
        CANONICAL_OWNER_UPN, CANONICAL_OWNER_ID,
        CANONICAL_DATE_CREATED,
    }
    _seen_canon: dict[str, str] = {}  # canonical -> first raw header that mapped to it
    for raw_h, canon_h in zip(raw_headers, canonical_headers):
        if canon_h not in _DUPE_GUARD:
            continue
        if canon_h in _seen_canon:
            prev = _seen_canon[canon_h]
            if prev.strip().lower() != raw_h.strip().lower():
                raise ValueError(
                    f"Ambiguous canonical header mapping: two distinct raw headers "
                    f"('{prev}', '{raw_h}') both map to canonical field '{canon_h}'. "
                    "Resolve the ambiguity in the column-map JSON (header_aliases)."
                )
        else:
            _seen_canon[canon_h] = raw_h

    # Hard-fail with header-diff if required canonical fields are missing.
    missing = required_fields - mapped_header_set
    if missing:
        raise ValueError(
            f"Registry export is missing required fields after header mapping: "
            f"{sorted(missing)}.\n"
            f"  Observed raw headers:       {raw_headers}\n"
            f"  Observed canonical headers: {sorted(mapped_header_set)}\n"
            f"Update the column-map JSON (header_aliases) if the export format changed."
        )

    rows: list[dict] = []
    warnings: list[str] = []
    for row_idx, raw_row in enumerate(all_rows[1:], start=2):
        # Pad or truncate row to match header count.
        padded = list(raw_row) + [None] * max(0, len(canonical_headers) - len(raw_row))

        # Skip fully-blank rows (every cell is None or whitespace-only).
        if all(
            v is None or (isinstance(v, str) and v.strip() == "")
            for v in padded[: len(canonical_headers)]
        ):
            logger.debug("Skipping fully-blank row at data row %d.", row_idx)
            continue

        row_raw: dict = {
            canonical_headers[i]: (
                str(padded[i]).strip()
                if padded[i] is not None and str(padded[i]).strip() != ""
                else None
            )
            for i in range(len(canonical_headers))
        }

        # Map canonical names to Dataverse logical names.
        # date_created is validated/normalised separately below.
        # owner_upn is hygiene-checked separately below — the hygiene block is
        # the SOLE writer of fsi_ownerupn so that a display-name-only value
        # (no '@') is never stored as a UPN.
        row_dv: dict = {}
        for canonical, dv_name in CANONICAL_TO_DATAVERSE.items():
            if canonical in (CANONICAL_DATE_CREATED, CANONICAL_OWNER_UPN):
                continue  # handled separately
            val = row_raw.get(canonical)
            if val is not None:
                row_dv[dv_name] = val

        # date_created: validate/normalise ISO-8601 UTC; invalid -> null + per-row warning.
        raw_date = row_raw.get(CANONICAL_DATE_CREATED)
        if raw_date is not None:
            try:
                row_dv["fsi_createdon"] = _parse_iso8601_utc(raw_date)
            except ValueError:
                warn_msg = (
                    f"Row {row_idx}: date_created value {raw_date!r} is not valid "
                    f"ISO-8601 UTC — fsi_createdon set to null for this row."
                )
                warnings.append(warn_msg)
                logger.warning(warn_msg)
                # fsi_createdon intentionally left absent (null).

        # Owner attribution provenance.
        owner_upn_raw = row_raw.get(CANONICAL_OWNER_UPN)
        owner_id = row_raw.get(CANONICAL_OWNER_ID)

        # owner_upn hygiene: a value without '@' is NOT a UPN (could be a display
        # name). Treat as absent — never store a display name as fsi_ownerupn.
        owner_upn: Optional[str] = None
        if owner_upn_raw and "@" in owner_upn_raw:
            owner_upn = owner_upn_raw
        elif owner_upn_raw:
            logger.debug(
                "Row %d: owner_upn %r lacks '@' — not UPN-shaped; "
                "treating as absent (no fsi_ownerupn set).",
                row_idx, owner_upn_raw,
            )

        if owner_upn is not None:
            row_dv["fsi_ownerupn"] = owner_upn

        row_dv["fsi_ownersource"] = OWNER_SOURCE_LABEL
        if owner_id:
            # Stable Entra object GUID present — highest confidence.
            confidence = "Exact"
        elif owner_upn is not None:
            # UPN only (no stable object ID) — heuristic match.
            confidence = "Heuristic"
        else:
            # Neither valid UPN nor owner_id: unresolved; not entitlement-classified.
            confidence = "Unmatched"
        row_dv["fsi_ownermatchconfidence"] = confidence
        if as_of:
            row_dv["fsi_ownerasofdatetime"] = as_of

        rows.append(row_dv)

    logger.info(
        "Parsed %d data rows from %s (%d per-row warnings).",
        len(rows), path, len(warnings),
    )
    return rows, warnings


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def _main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--input", required=True,
        help="Path to the registry export file (.xlsx or .csv).",
    )
    parser.add_argument(
        "--columnmap",
        default=str(_DEFAULT_COLUMNMAP),
        help=(
            "Path to the column-map JSON (header alias map). "
            f"Default: {_DEFAULT_COLUMNMAP}"
        ),
    )
    parser.add_argument(
        "--as-of",
        default=None,
        help=(
            "ISO-8601 UTC datetime representing when the export was produced "
            "(e.g., 2026-07-20T18:00:00Z). Stamped as fsi_ownerasofdatetime on "
            "every row to signal export staleness. Invalid values are rejected."
        ),
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Write the enriched row JSON to this path (default: stdout).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and validate the file without writing output.",
    )
    parser.add_argument(
        "--log-level",
        default=os.environ.get("CAI_LOG_LEVEL", "INFO"),
        help="Logging level (DEBUG, INFO, WARNING, ERROR).",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=getattr(logging, str(args.log_level).upper(), logging.INFO),
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )

    # Validate --as-of before doing any file I/O.
    as_of_normalized: Optional[str] = None
    if args.as_of:
        try:
            as_of_normalized = _parse_iso8601_utc(args.as_of)
        except ValueError as exc:
            logger.error("--as-of invalid: %s", exc)
            return 1

    columnmap_path = Path(args.columnmap)
    aliases, required_fields, sheet_name = load_columnmap(columnmap_path)

    file_hash = sha256_file(args.input)
    logger.info(
        "Registry export: %s (sha256=%s...)",
        args.input, file_hash[:12],
    )

    try:
        rows, import_warnings = import_registry_file(
            path=args.input,
            aliases=aliases,
            required_fields=required_fields,
            sheet_name=sheet_name,
            as_of=as_of_normalized,
        )
    except ValueError as exc:
        logger.error("Import failed (header validation): %s", exc)
        return 1

    # Stamp provenance on the result envelope.
    result = {
        "source_file": os.path.basename(args.input),
        "source_sha256": file_hash,
        "row_count": len(rows),
        "as_of": as_of_normalized,
        "warnings": import_warnings,
        "rows": rows,
    }

    if args.dry_run:
        logger.info(
            "[DRY RUN] would emit %d enriched rows (%d warnings, no output written). "
            "First row preview: %s",
            len(rows), len(import_warnings),
            json.dumps(rows[0], default=str) if rows else "{}",
        )
        return 0

    output_json = json.dumps(result, indent=2, default=str)
    if args.output:
        Path(args.output).write_text(output_json, encoding="utf-8")
        logger.info("Wrote %d enriched rows to %s", len(rows), args.output)
    else:
        sys.stdout.write(output_json + "\n")

    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
